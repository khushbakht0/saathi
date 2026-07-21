from __future__ import annotations

import logging
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.core.logger import logger

class ExcelWorkbookParser:
    """Parse workbook-based timetable sources and extract structured rows."""

    REQUIRED_HEADERS = {
        "day",
        "time",
        "course_code",
        "course_name",
        "section",
        "faculty",
        "room",
        "building",
    }

    def parse(self, workbook_path: str | Path) -> list[dict[str, Any]]:
        workbook = Path(workbook_path)
        logger.info("starting workbook parse", extra={"path": str(workbook)})

        try:
            wb = load_workbook(workbook, data_only=True)
        except Exception as exc:
            logger.exception("failed to load workbook for parsing")
            raise ValueError(f"Unable to read workbook: {exc}") from exc

        records: list[dict[str, Any]] = []

        for sheet in wb.worksheets:
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                continue

            header_match = self._find_header_row(rows)
            if header_match is None:
                continue

            header_row_index, header_row = header_match
            headers = [self._normalize_header(str(cell)).strip() for cell in header_row]
            column_map = self._build_column_map(headers)
            if not column_map:
                continue

            for row in rows[header_row_index + 1 :]:
                if not any(cell is not None and str(cell).strip() for cell in row):
                    continue
                record = self._build_record(row, column_map)
                if record is not None:
                    records.append(record)

        logger.info("workbook parsing completed", extra={"records_found": len(records)})
        return records

    def _find_header_row(self, rows: list[tuple[Any, ...]]) -> tuple[int, tuple[Any, ...]] | None:
        for index, row in enumerate(rows):
            normalized = [self._normalize_header(str(cell)) for cell in row]
            if set(self.REQUIRED_HEADERS).issubset(set(normalized)):
                return index, row
        return None

    def _build_column_map(self, headers: list[str]) -> dict[str, int]:
        lookup: dict[str, int] = {}
        for index, header in enumerate(headers):
            normalized = self._normalize_header(header)
            if normalized in self.REQUIRED_HEADERS:
                lookup[normalized] = index
        return lookup

    def _build_record(self, row: tuple[Any, ...], column_map: dict[str, int]) -> dict[str, Any] | None:
        try:
            day = self._coerce_value(row, column_map, "day")
            time_range = self._coerce_value(row, column_map, "time")
            course_code = self._coerce_value(row, column_map, "course_code")
            course_name = self._coerce_value(row, column_map, "course_name")
            section = self._coerce_value(row, column_map, "section")
            faculty = self._coerce_value(row, column_map, "faculty")
            room = self._coerce_value(row, column_map, "room")
            building = self._coerce_value(row, column_map, "building")
        except KeyError:
            logger.warning("skipping malformed row due to missing mapped columns")
            return None

        if not all([day, time_range, course_code, course_name, section, faculty, room]):
            logger.warning("skipping incomplete row", extra={"row": row})
            return None

        return {
            "day": day,
            "time": time_range,
            "course_code": course_code,
            "course_name": course_name,
            "section": section,
            "faculty": faculty,
            "room": room,
            "building": building,
            "batch": self._coerce_value(row, column_map, "batch", optional=True),
        }

    def _coerce_value(self, row: tuple[Any, ...], column_map: dict[str, int], key: str, optional: bool = False) -> str | None:
        if key not in column_map:
            if optional:
                return None
            raise KeyError(key)

        value = row[column_map[key]]
        if value is None:
            if optional:
                return None
            raise KeyError(key)
        return str(value).strip()

    def _normalize_header(self, header: str) -> str:
        header = header.lower().replace(" ", "_").replace("-", "_")
        return header.replace("(", "").replace(")", "")
